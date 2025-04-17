import logging
import os
import tempfile
from copy import copy
from datetime import date
from typing import Optional

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

from config import settings
from entities import Report

logger = logging.getLogger(__name__)

# Номер строчки (с 1) которая используется как шаблон стилей
TEMPLATE_ROW_IDX = 2
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def create_excel_report_file(report_on: date, reports: list[Report]) -> str:
    wb, _ = create_excel_report(report_on, reports)
    filename = f"{report_on.strftime('%d%m%Y')}.xlsx"
    file_path, = os.path.join(tempfile.gettempdir(), filename)
    save_excel(wb, file_path)
    logger.info(f"Created report on '{report_on}': {file_path}")
    return file_path


def create_excel_report(report_on, reports) -> tuple[Workbook, int]:
    wb = create_excel_from_template(report_on)
    next_row_idx = append_reports_to_excel(wb, TEMPLATE_ROW_IDX, reports)
    return wb, next_row_idx


def create_excel_from_template(report_on):
    wb = load_workbook(settings.report_template_path)
    ws = wb.active
    ws.title = ws.title.format(report_on=report_on.strftime("%d.%m.%Y"))
    return wb


def append_reports_to_excel(wb: Workbook, start_row_idx: int, reports: list[Report]) -> int:
    ws = wb.active
    for i, report in enumerate(reports):
        row_idx = start_row_idx + i
        vals = [
            report.worked_on,
            resolve_dict_value(
                report.department.subdivision if report.department else None,
                report.department_raw,
                report.department_predicted
            ),
            resolve_dict_value(
                report.operation.operation_name if report.operation else None,
                report.operation_raw,
                report.operation_predicted
            ),
            resolve_dict_value(
                report.crop.crop_name if report.crop else None,
                report.crop_raw,
                report.crop_predicted
            ),
            report.day_area,
            report.cumulative_area,
            report.day_yield or None,
            report.cumulative_yield or None
        ]
        for col_idx, value in enumerate(vals, start=1):
            tmp = ws.cell(row=TEMPLATE_ROW_IDX, column=col_idx)
            dst = ws.cell(row=row_idx, column=col_idx, value=value)
            dst.font = copy(tmp.font)
            dst.border = copy(tmp.border)
            dst.fill = copy(tmp.fill)
            dst.protection = copy(tmp.protection)
            dst.alignment = copy(tmp.alignment)
            dst.number_format = copy(tmp.number_format)

            if col_idx in (2, 3, 4):
                if (
                        (col_idx == 2 and not report.department) or
                        (col_idx == 3 and not report.operation) or
                        (col_idx == 4 and not report.crop)
                ):
                    dst.fill = YELLOW_FILL

    next_row_idx = start_row_idx + len(reports)
    return next_row_idx


def resolve_dict_value(
        dict_value: Optional[str],
        raw_value: Optional[str],
        predicted_value: Optional[str]
) -> Optional[str]:
    if dict_value:
        return dict_value
    if raw_value:
        return raw_value
    if predicted_value:
        return predicted_value
    return None


def save_excel(wb: Workbook, file_path: str) -> str:
    with open(file_path, "w+t"):
        wb.save(file_path)
        return file_path
